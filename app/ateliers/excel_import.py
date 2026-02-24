from __future__ import annotations

import re
from datetime import datetime, date
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

from app.extensions import db
from app.models import AtelierActivite, SessionActivite, PresenceActivite, Participant, Quartier


# ------------------------------------------------------------
# Config
# ------------------------------------------------------------
SKIP_SHEETS_DEFAULT = {
    "feuil1",
    "sheet1",
    "adultes",
    "adultes saveurs",
}

FOOTER_ROW_PREFIXES = (
    "NOMBRE",
    "NBRE",
    "TAUX",
    "MOYENNE",
    "DUREE",
    "DURÉE",
    "CAPACITE",
    "CAPACITÉ",
    "TOTAL",
)

CREIL_CANON = "Creil"


# ------------------------------------------------------------
# Helpers robustes
# ------------------------------------------------------------
def _norm(s: Any) -> str:
    return (str(s or "")).strip()


def _norm_key(s: Any) -> str:
    """
    Normalisation agressive pour comparaisons (headers, villes, etc.).
    - minuscule
    - enlève (60)
    - remplace ponctuation par espaces
    - compacte les espaces
    """
    x = _norm(s).lower()
    x = x.replace("(60)", "")
    x = x.replace("’", "'")
    x = x.replace("'", " ").replace("-", " ").replace("_", " ").replace("/", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return x


def normalize_secteur(value: Any) -> str:
    v = str(value or "").strip()
    if not v:
        return "Numérique"
    low = _norm_key(v)
    if low in ("numerique", "numérique", "num", "nume", "numeriq", "numerik"):
        return "Numérique"
    return v


def normalize_ville(value: Any) -> Optional[str]:
    v = _norm(value)
    if not v:
        return None
    v = v.replace("(60)", "")
    v = re.sub(r"\s+", " ", v).strip()

    # Quelques homogénéisations courantes (optionnel)
    # ex: "Nogent sur oise" => "Nogent Sur Oise"
    return v[:120].title()


def is_creil(ville: Optional[str]) -> bool:
    if not ville:
        return False
    return _norm_key(ville) == _norm_key(CREIL_CANON)


def normalize_genre(value: Any) -> Optional[str]:
    if value is None:
        return None
    v = str(value).strip().lower()
    if not v:
        return None

    # enlève ponctuation/espaces/char spéciaux
    v = re.sub(r"[^a-zàâäéèêëîïôöùûüç]", "", v)

    if v in ("f", "femme", "féminin", "feminin", "fille"):
        return "Femme"
    if v in ("h", "homme", "m", "masculin", "garcon", "garçon", "g"):
        return "Homme"
    return None


def _is_presence(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return bool(v)
    if isinstance(v, (int, float)):
        return v > 0
    s = _norm(v).lower()
    return s in {"1", "x", "p", "présent", "present", "oui", "o"}


def _to_date(v: Any) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue

    return None


def _is_header_cell(v: Any) -> bool:
    up = _norm(v).upper()
    return up in ("NOMS", "NOM") or up.startswith("NOMS")


def _is_prenom_cell(v: Any) -> bool:
    up = _norm(v).upper()
    return up.startswith("PRENOM") or up.startswith("PRÉNOM")


def _header_map(ws, header_row: int) -> Dict[str, int]:
    """
    Mapping "champ -> colonne" robuste.
    Tolère : espaces, slash, variantes ("COMMUNE", "VILLE/COMMUNE", etc.)
    """
    mapping: Dict[str, int] = {}

    tokens = {
        "ANNEE": ["annee", "année", "annee naissance", "année naissance", "annee de naissance", "naissance", "ddn", "date naissance"],
        "SEXE": ["sexe", "genre"],
        "QUARTIER": ["quartier", "quartiers"],
        "VILLE": ["ville", "commune"],
        "EMAIL": ["email", "e mail", "mail"],
        "TEL": ["tel", "tél", "telephone", "téléphone"],
    }

    for c in range(1, ws.max_column + 1):
        hv = _norm_key(ws.cell(header_row, c).value)
        if not hv:
            continue

        for key, want in tokens.items():
            if any(w in hv for w in want):
                mapping.setdefault(key, c)  # n'écrase pas si déjà trouvé

    return mapping


def _detect_date_columns(ws, header_row: int) -> Dict[int, date]:
    """
    Blindé: détecte des dates dans une zone autour de header_row.
    Retourne dict: col_index -> date.
    """
    date_by_col: Dict[int, date] = {}

    r_min = max(1, header_row - 4)
    r_max = min(ws.max_row, header_row + 10)

    for r in range(r_min, r_max + 1):
        for c in range(1, ws.max_column + 1):
            d = _to_date(ws.cell(r, c).value)
            if d and c not in date_by_col:
                date_by_col[c] = d

    return date_by_col


def _get_or_create_quartier(nom_quartier: str, ville: str = CREIL_CANON) -> Optional[Quartier]:
    """
    ⚠️ Dans ta stratégie : on ne crée/associe des quartiers que pour Creil.
    """
    nom = _norm(nom_quartier)
    if not nom:
        return None

    ville = normalize_ville(ville) or CREIL_CANON

    q = Quartier.query.filter_by(ville=ville, nom=nom).first()
    if q:
        return q

    up = nom.upper()
    is_qpv = ("ROUHER" in up) or ("HAUT" in up and "CREIL" in up) or ("QPV" in up)

    q = Quartier(ville=ville, nom=nom, is_qpv=is_qpv)
    db.session.add(q)

    try:
        with db.session.begin_nested():
            db.session.flush()
    except IntegrityError:
        return Quartier.query.filter_by(ville=ville, nom=nom).first()

    return q


def _find_participant(
    nom: str,
    prenom: str,
    year_or_ddn: Any,
    sexe: Any,
    quartier: Optional[Quartier],
    secteur: str,
    ville: Optional[str] = None,
) -> Participant:
    n = _norm(nom)
    p = _norm(prenom)

    # année -> date_naissance 01/01/année si plausible
    ddn: Optional[date] = None
    y = None
    try:
        y = int(str(year_or_ddn).strip())
    except Exception:
        y = None
    if y and 1900 <= y <= 2030:
        ddn = date(y, 1, 1)

    q = Participant.query.filter(Participant.nom.ilike(n), Participant.prenom.ilike(p))
    if ddn:
        q = q.filter(Participant.date_naissance == ddn)

    existing = q.first()

    nv = normalize_ville(ville) if ville else None
    ng = normalize_genre(sexe)

    if existing:
        changed = False

        # ville : on la met si vide (ou placeholder) — ET on n’écrase pas une ville déjà renseignée
        if nv and getattr(existing, "ville", None) in (None, "", "—"):
            existing.ville = nv
            changed = True

        # quartier seulement si Creil et si manquant
        if quartier and is_creil(nv) and not existing.quartier_id:
            existing.quartier = quartier
            changed = True

        if ddn and not existing.date_naissance:
            existing.date_naissance = ddn
            changed = True

        if ng and existing.genre != ng:
            existing.genre = ng
            changed = True

        if changed:
            db.session.add(existing)

        return existing

    part = Participant(
        nom=n or "?",
        prenom=p or "?",
        genre=ng,
        date_naissance=ddn,
        quartier=quartier if is_creil(nv) else None,  # ✅ NULL hors Creil
        created_secteur=secteur,
    )

    if nv is not None and hasattr(part, "ville"):
        part.ville = nv

    db.session.add(part)
    db.session.flush()
    return part


def _get_or_create_atelier(secteur: str, nom_atelier: str) -> AtelierActivite:
    nom = _norm(nom_atelier)
    a = AtelierActivite.query.filter_by(secteur=secteur, nom=nom, is_deleted=False).first()
    if a:
        return a
    a = AtelierActivite(secteur=secteur, nom=nom, type_atelier="COLLECTIF")
    db.session.add(a)
    db.session.flush()
    return a


def _get_or_create_session(atelier: AtelierActivite, secteur: str, d: date) -> SessionActivite:
    s = (
        SessionActivite.query.filter_by(
            atelier_id=atelier.id,
            secteur=secteur,
            session_type="COLLECTIF",
            date_session=d,
            is_deleted=False,
        )
        .first()
    )
    if s:
        return s

    s = SessionActivite(
        atelier_id=atelier.id,
        secteur=secteur,
        session_type="COLLECTIF",
        date_session=d,
        statut="realisee",
    )
    db.session.add(s)
    db.session.flush()
    return s


# ------------------------------------------------------------
# Import principal
# ------------------------------------------------------------
def import_presences_from_xlsx(
    xlsx_path: str,
    secteur: str = "NUMERIQUE",
    skip_sheets: Optional[List[str]] = None,
    dry_run: bool = False,
    limit_rows_per_sheet: int = 5000,
) -> Dict[str, Any]:

    secteur = normalize_secteur(secteur)

    stats: Dict[str, Any] = {
        "ateliers_created": 0,
        "sessions_created": 0,
        "participants_created": 0,
        "presences_created": 0,
        "presences_skipped_duplicates": 0,
        "sheets_processed": 0,
        "warnings": [],
        "sheets": {},
    }

    skip = {s.strip().lower() for s in (skip_sheets or [])}
    if not skip_sheets:
        skip = set(SKIP_SHEETS_DEFAULT)

    wb = load_workbook(xlsx_path, data_only=True)

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in skip:
                continue

            ws = wb[sheet_name]

            # 1) Trouver header_row (NOMS / PRENOMS)
            header_row = None
            for r in range(1, min(100, ws.max_row) + 1):
                v1 = ws.cell(r, 1).value
                v2 = ws.cell(r, 2).value
                if _is_header_cell(v1) and _is_prenom_cell(v2):
                    header_row = r
                    break

            if not header_row:
                stats["warnings"].append(f"Onglet '{sheet_name}': entête NOMS/PRENOMS introuvable -> ignoré.")
                continue

            # 2) Mapping colonnes par en-têtes (robuste)
            colmap = _header_map(ws, header_row)

            col_annee = colmap.get("ANNEE", 3)
            col_sexe = colmap.get("SEXE", 5)
            col_quart = colmap.get("QUARTIER", 6)
            col_ville = colmap.get("VILLE", None)

            # 3) Détection dates blindée
            date_by_col = _detect_date_columns(ws, header_row)
            date_cols = sorted(date_by_col.keys())

            # garde les dates "à droite" des identités
            date_cols = [c for c in date_cols if c >= 3]

            if not date_cols:
                stats["warnings"].append(f"Onglet '{sheet_name}': aucune colonne date détectée -> ignoré.")
                continue

            stats_sheet: Dict[str, Any] = {
                "sheet": sheet_name,
                "participants": 0,
                "presences": 0,
                "sessions": 0,
                "warnings": [],
                "colmap": {
                    "ANNEE": col_annee,
                    "SEXE": col_sexe,
                    "QUARTIER": col_quart,
                    "VILLE": col_ville,
                },
                "dates_detectees": len(date_cols),
            }

            if col_ville is None:
                stats_sheet["warnings"].append(
                    "⚠️ Colonne VILLE non détectée (header_map). "
                    "Fallback ville = Creil. Quartiers hors Creil resteront NULL."
                )

            atelier = _get_or_create_atelier(secteur, sheet_name.strip())

            blank_streak = 0
            processed_rows = 0

            for r in range(header_row + 1, ws.max_row + 1):
                if processed_rows >= limit_rows_per_sheet:
                    stats_sheet["warnings"].append(f"Limite de lignes atteinte ({limit_rows_per_sheet}).")
                    break

                nom = ws.cell(r, 1).value
                prenom = ws.cell(r, 2).value

                nom_s = _norm(nom)
                prenom_s = _norm(prenom)

                # vides
                if (not nom_s) and (not prenom_s):
                    blank_streak += 1
                    if blank_streak >= 8:
                        break
                    continue
                blank_streak = 0

                # stop footer
                up_nom = nom_s.upper()
                if (not prenom_s) and up_nom.startswith(FOOTER_ROW_PREFIXES):
                    break
                if up_nom == "TOTAL" or up_nom.startswith("TOTAL"):
                    break
                if nom_s and not prenom_s:
                    continue

                processed_rows += 1

                annee_naissance = ws.cell(r, col_annee).value if col_annee else None
                sexe = ws.cell(r, col_sexe).value if col_sexe else None
                quartier_name = ws.cell(r, col_quart).value if col_quart else None

                # Ville : IMPORTANT
                # - si col_ville détectée : on lit cellule, sinon fallback Creil
                # - si cellule vide MAIS col_ville existe : on met None (évite coller Creil par erreur)
                if col_ville is None:
                    ville_norm = CREIL_CANON
                else:
                    ville_norm = normalize_ville(ws.cell(r, col_ville).value) or None

                # si tu veux absolument une valeur : fallback Creil
                # (tu peux commenter si tu préfères des villes NULL quand cellule vide)
                ville_norm = ville_norm or CREIL_CANON

                # Quartier: NULL hors Creil
                quartier = None
                if is_creil(ville_norm):
                    qn = _norm(quartier_name)
                    if qn:
                        quartier = _get_or_create_quartier(qn, ville=CREIL_CANON)

                # déterminer si "nouveau" (compteur)
                before = Participant.query.filter(
                    Participant.nom.ilike(_norm(nom)),
                    Participant.prenom.ilike(_norm(prenom)),
                ).first()

                part = _find_participant(
                    _norm(nom),
                    _norm(prenom),
                    annee_naissance,
                    sexe,
                    quartier,
                    secteur,
                    ville=ville_norm,
                )

                if before is None:
                    stats["participants_created"] += 1
                stats_sheet["participants"] += 1

                # présences
                for c in date_cols:
                    d = date_by_col.get(c)
                    if not d:
                        continue

                    v = ws.cell(r, c).value
                    if not _is_presence(v):
                        continue

                    session_before = SessionActivite.query.filter_by(
                        atelier_id=atelier.id,
                        secteur=secteur,
                        session_type="COLLECTIF",
                        date_session=d,
                        is_deleted=False,
                    ).first()

                    session = _get_or_create_session(atelier, secteur, d)
                    if session_before is None:
                        stats["sessions_created"] += 1
                        stats_sheet["sessions"] += 1

                    pr = PresenceActivite(session_id=session.id, participant_id=part.id)
                    try:
                        with db.session.begin_nested():
                            db.session.add(pr)
                            db.session.flush()
                        stats["presences_created"] += 1
                        stats_sheet["presences"] += 1
                    except IntegrityError:
                        stats["presences_skipped_duplicates"] += 1

            stats["sheets_processed"] += 1
            stats["sheets"][sheet_name] = stats_sheet

        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()

    except Exception:
        db.session.rollback()
        raise

    return stats
