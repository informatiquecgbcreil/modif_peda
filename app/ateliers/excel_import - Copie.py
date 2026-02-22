from __future__ import annotations

import tempfile
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

from app.extensions import db
from app.models import AtelierActivite, SessionActivite, PresenceActivite, Participant, Quartier


SKIP_SHEETS_DEFAULT = {
    "feuil1",
    "sheet1",
    "adultes",
    "adultes saveurs",
}


# Lignes "métier" en bas des tableaux qui ne sont PAS des participants.
# On s'arrête dès qu'on tombe dessus, sinon on crée des faux participants du genre "CAPACITE D'ACCUEIL ?".
FOOTER_ROW_PREFIXES = (
    "NOMBRE",
    "NBRE",
    "TAUX",
    "MOYENNE",
    "DUREE",
    "DURÉE",
    "CAPACITE",
    "CAPACITÉ",
    "TOTAL",
)


def _norm(s: Any) -> str:
    return (str(s or "")).strip()
    
def normalize_genre(value: Any) -> Optional[str]:
    if value is None:
        return None
    v = str(value).strip().lower()
    if not v:
        return None
    if v in ("f", "Fille", "fille", "femme"):
        return "Femme"
    if v in ("h", "m", "g", "garçon", "Garçon", "garcon", "Garcon", "homme"):
        return "Homme"
    return None


def normalize_secteur(value: Any) -> str:
    v = str(value or "").strip()
    if not v:
        return "Numérique"
    low = v.lower()
    if low in ("numerique", "numérique", "num", "nume"):
        return "Numérique"
    return v


def _is_presence(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return bool(v)
    if isinstance(v, (int, float)):
        return v > 0
    s = _norm(v).lower()
    return s in {"1", "x", "p", "présent", "present", "oui", "o"}


def _to_date(v: Any) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None

        # formats fréquents Excel FR
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue

    return None


def _get_or_create_quartier(nom_quartier: str, ville: str = "Creil") -> Optional[Quartier]:
    nom = _norm(nom_quartier)
    if not nom:
        return None

    q = Quartier.query.filter_by(ville=ville, nom=nom).first()
    if q:
        return q

    # Heuristique QPV (Creil)
    up = nom.upper()
    is_qpv = ("ROUHER" in up) or ("HAUT" in up and "CREIL" in up) or ("QPV" in up)

    q = Quartier(ville=ville, nom=nom, is_qpv=is_qpv)
    db.session.add(q)
    # Utiliser un SAVEPOINT pour éviter de rollback toute la session sur collision d'unicité
    try:
        with db.session.begin_nested():
            db.session.flush()
    except IntegrityError:
        # Collision probable (ex: import concurrent) -> on récupère l'existant
        return Quartier.query.filter_by(ville=ville, nom=nom).first()
    return q


def _find_participant(nom: str, prenom: str, year_or_ddn: Any, sexe: Any, quartier: Optional[Quartier], secteur: str) -> Participant:
    n = _norm(nom)
    p = _norm(prenom)

    # Essai 1: match strict nom/prenom (+date_naissance si possible)
    ddn: Optional[date] = None
    y = None
    try:
        y = int(str(year_or_ddn).strip())
    except Exception:
        y = None
    if y and 1900 <= y <= 2030:
        ddn = date(y, 1, 1)

    q = Participant.query.filter(Participant.nom.ilike(n), Participant.prenom.ilike(p))
    if ddn:
        q = q.filter(Participant.date_naissance == ddn)
    existing = q.first()
    if existing:
        # On complète si des infos manquent
        changed = False
        if quartier and not existing.quartier_id:
            existing.quartier = quartier
            changed = True
        if ddn and not existing.date_naissance:
            existing.date_naissance = ddn
            changed = True
        new_genre = normalize_genre(sexe)
        if new_genre and existing.genre != new_genre:
            existing.genre = new_genre
            changed = True
        if changed:
            db.session.add(existing)
        return existing

    part = Participant(
        nom=n or "?",
        prenom=p or "?",
        genre=normalize_genre(sexe),
        date_naissance=ddn,
        quartier=quartier,
        created_secteur=secteur,
    )
    db.session.add(part)
    db.session.flush()
    return part


def _get_or_create_atelier(secteur: str, nom_atelier: str) -> AtelierActivite:
    nom = _norm(nom_atelier)
    a = AtelierActivite.query.filter_by(secteur=secteur, nom=nom, is_deleted=False).first()
    if a:
        return a
    a = AtelierActivite(secteur=secteur, nom=nom, type_atelier="COLLECTIF")
    db.session.add(a)
    db.session.flush()
    return a


def _get_or_create_session(atelier: AtelierActivite, secteur: str, d: date) -> SessionActivite:
    s = (
        SessionActivite.query.filter_by(
            atelier_id=atelier.id,
            secteur=secteur,
            session_type="COLLECTIF",
            date_session=d,
            is_deleted=False,
        )
        .first()
    )
    if s:
        return s
    s = SessionActivite(
        atelier_id=atelier.id,
        secteur=secteur,
        session_type="COLLECTIF",
        date_session=d,
        statut="realisee",
    )
    db.session.add(s)
    db.session.flush()
    return s


def import_presences_from_xlsx(
    xlsx_path: str,
    secteur: str = "NUMERIQUE",
    skip_sheets: Optional[List[str]] = None,
    dry_run: bool = False,
    limit_rows_per_sheet: int = 5000,
) -> Dict[str, Any]:
    """Importe un fichier Excel "présences" (format Antoine).

    Retourne un dict de stats + warnings.
    - Crée les Ateliers (par onglet)
    - Crée les Sessions (par date de colonne)
    - Crée les Participants (par ligne)
    - Crée les Présences (cellule non vide)

    ⚠️ Heures / lieux / intervenants: pas dispo dans ce format -> sessions créées "date-only".
    """
    stats = {
        "ateliers_created": 0,
        "sessions_created": 0,
        "participants_created": 0,
        "presences_created": 0,
        "presences_skipped_duplicates": 0,
        "sheets_processed": 0,
        "warnings": [],
        "sheets": {},
    }

    skip = {s.strip().lower() for s in (skip_sheets or [])}
    if not skip_sheets:
        skip = set(SKIP_SHEETS_DEFAULT)

    wb = load_workbook(xlsx_path, data_only=True)

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in skip:
                continue

            ws = wb[sheet_name]
            # Trouver la ligne d'entête (NOMS / PRENOMS)
            header_row = None
            for r in range(1, min(60, ws.max_row) + 1):
                v1 = ws.cell(r, 1).value
                v2 = ws.cell(r, 2).value
                if _norm(v1).upper() == "NOMS" and _norm(v2).upper().startswith("PRENOM"):
                    header_row = r
                    break

            if not header_row:
                stats["warnings"].append(f"Onglet '{sheet_name}': entête NOMS/PRENOMS introuvable -> ignoré.")
                continue

            # Colonnes dates
            date_cols: List[int] = []
            for c in range(1, ws.max_column + 1):
                hv = ws.cell(header_row, c).value
                if _to_date(hv):
                    date_cols.append(c)

            if not date_cols:
                stats["warnings"].append(f"Onglet '{sheet_name}': aucune colonne date détectée -> ignoré.")
                continue

            atelier = _get_or_create_atelier(secteur, sheet_name.strip())
            # Detect atelier created via transient flush (no easy way) -> count by existence check before
            stats_sheet = {
                "sheet": sheet_name,
                "participants": 0,
                "presences": 0,
                "sessions": 0,
                "warnings": [],
            }

            blank_streak = 0
            processed_rows = 0

            for r in range(header_row + 1, ws.max_row + 1):
                if processed_rows >= limit_rows_per_sheet:
                    stats_sheet["warnings"].append(f"Limite de lignes atteinte ({limit_rows_per_sheet}).")
                    break

                nom = ws.cell(r, 1).value
                prenom = ws.cell(r, 2).value

                nom_s = _norm(nom)
                prenom_s = _norm(prenom)

                # Lignes vides -> compteur de streak
                if (not nom_s) and (not prenom_s):
                    blank_streak += 1
                    if blank_streak >= 8:
                        break
                    continue
                blank_streak = 0
                
                # Stop sur les lignes de stats métier (en bas du tableau)
                up_nom = nom_s.upper()
                if (not prenom_s) and up_nom.startswith(FOOTER_ROW_PREFIXES):
                    break
                # Ignore les lignes 'TOTAL' ou assimilées
                if up_nom in ("TOTAL",) or up_nom.startswith("TOTAL"):
                    break
                # Si prénom vide mais nom rempli, on ignore (évite faux participants)
                if nom_s and not prenom_s:
                    continue

                processed_rows += 1

                annee_naissance = ws.cell(r, 3).value
                sexe = ws.cell(r, 5).value
                quartier_name = ws.cell(r, 6).value

                quartier = _get_or_create_quartier(_norm(quartier_name)) if _norm(quartier_name) else None

                # participant
                # Determine if participant existed before to count creations
                before = Participant.query.filter(Participant.nom.ilike(_norm(nom)), Participant.prenom.ilike(_norm(prenom))).first()
                part = _find_participant(_norm(nom), _norm(prenom), annee_naissance, sexe, quartier, secteur)
                if before is None:
                    stats["participants_created"] += 1
                stats_sheet["participants"] += 1

                # presences
                for c in date_cols:
                    d = _to_date(ws.cell(header_row, c).value)
                    if not d:
                        continue
                    v = ws.cell(r, c).value
                    if not _is_presence(v):
                        continue

                    # session
                    session_before = SessionActivite.query.filter_by(
                        atelier_id=atelier.id, secteur=secteur, session_type="COLLECTIF", date_session=d, is_deleted=False
                    ).first()
                    session = _get_or_create_session(atelier, secteur, d)
                    if session_before is None:
                        stats["sessions_created"] += 1
                        stats_sheet["sessions"] += 1

                    pr = PresenceActivite(session_id=session.id, participant_id=part.id)
                    # SAVEPOINT par insertion pour éviter de rollback toute la session sur un doublon
                    try:
                        with db.session.begin_nested():
                            db.session.add(pr)
                            db.session.flush()
                        stats["presences_created"] += 1
                        stats_sheet["presences"] += 1
                    except IntegrityError:
                        stats["presences_skipped_duplicates"] += 1

            stats["sheets_processed"] += 1
            stats["sheets"][sheet_name] = stats_sheet

        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise

    return stats
